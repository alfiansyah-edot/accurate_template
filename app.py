import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from datetime import datetime

st.title("Accurate SO Template Generator")

uploaded_file = st.file_uploader("Upload your Sales Order file", type=["csv", "xlsx"])

if uploaded_file is not None:
    with st.spinner("Processing file, please wait..."):
        # Read file based on extension
        if uploaded_file.name.endswith(".csv"):
            data = pd.read_csv(uploaded_file)
        else:
            data = pd.read_excel(uploaded_file)

        # Filter and convert date
        data = data[data["type"] == "item"]
        data["date"] = pd.to_datetime(data["date"])

        # Ranking logic
        data = data.sort_values(by=["date", "number"])
        data["rank"] = data.drop_duplicates(subset=["date", "number"]).reset_index(drop=True).reset_index().set_index(["date", "number"]).reindex(data.set_index(["date", "number"]).index)["index"].values + 1
        data["product_rank"] = data.groupby("number")["product"].rank(method="dense").astype(int)

        data["date"] = data["date"].dt.strftime('%d/%m/%Y')
        data = data.sort_values(by=["rank", "product_rank"])

        # Prepare header, item, expense rows
        header_rows = data[["rank", "date", "customer_external_code","price_before_tax","price_after_tax","discount","branch","payment_terms"]].copy().drop_duplicates()
        header_rows["row_or_header"] = 2
        header_rows["rank_part"] = 1
        header_rows["header"] = "HEADER"
        header_rows["no_form"] = ""
        header_rows["tgl_pesanan"] = header_rows["date"]  # keep datetime
        header_rows["no_pelanggan"] = header_rows["customer_external_code"]  # keep original type
        header_rows["no_po"] = ""
        header_rows["alamat"] = ""
        header_rows["kena_ppn"]=(header_rows['price_before_tax'] != header_rows["price_after_tax"]).map({True: "Ya", False: "Tidak"})
        header_rows["total_termasuk_ppn"]=(header_rows["price_before_tax"] != header_rows["price_after_tax"]).map({True: "Ya", False: "Tidak"})
        header_rows["diskon_pesanan_percentage"] = ""
        header_rows["diskon_pesanan_rupiah"] = header_rows["discount"].astype(float).round(2)
        header_rows["keterangan"] = ""
        header_rows["nama_cabang"] = ""
        header_rows["pengiriman"] = ""
        header_rows["tanggal_pengiriman"] = ""
        header_rows["FOB"] = ""
        header_rows["syarat_pembayaran"] = ""
        header_rows = header_rows.groupby(
            ["row_or_header", "rank", "rank_part", "header", "no_form", "tgl_pesanan", "no_pelanggan","no_po","alamat","kena_ppn","total_termasuk_ppn","diskon_pesanan_percentage","keterangan","nama_cabang","pengiriman","tanggal_pengiriman","FOB","syarat_pembayaran"],
            as_index=False
        ).agg({'diskon_pesanan_rupiah': 'sum'})
        header_rows = header_rows[["row_or_header", "rank", "rank_part", "header", "no_form", "tgl_pesanan", "no_pelanggan","no_po","alamat","kena_ppn","total_termasuk_ppn","diskon_pesanan_percentage","diskon_pesanan_rupiah","keterangan","nama_cabang","pengiriman","tanggal_pengiriman","FOB","syarat_pembayaran"]]
        
        # ITEM rows
        item_rows = data[["rank", "product_external_code", "product", "qty","uom","qty_price","salesman_external_code"]].copy()
        item_rows["row_or_header"] = 2
        item_rows["rank_part"] = 2
        item_rows["header"] = "ITEM"
        item_rows["no_form"] = item_rows["product_external_code"]
        item_rows["tgl_pesanan"] = item_rows["product"]
        item_rows["no_pelanggan"] = pd.to_numeric(item_rows["qty"], errors='coerce')
        item_rows["no_po"] = item_rows["uom"]
        item_rows["alamat"] = item_rows["qty_price"].astype(float).round(2)
        item_rows["kena_ppn"] = ""
        item_rows["total_termasuk_ppn"] = ""
        item_rows["diskon_pesanan_percentage"] = ""
        item_rows["diskon_pesanan_rupiah"] = ""
        item_rows["keterangan"] = ""
        item_rows["nama_cabang"] = ""
        item_rows["pengiriman"] = item_rows["salesman_external_code"]
        item_rows["tanggal_pengiriman"] = ""
        item_rows["FOB"] = ""
        item_rows["syarat_pembayaran"] = ""
        item_rows = item_rows[["row_or_header", "rank", "rank_part", "header", "no_form", "tgl_pesanan", "no_pelanggan","no_po","alamat","kena_ppn","total_termasuk_ppn","diskon_pesanan_percentage","diskon_pesanan_rupiah","keterangan","nama_cabang","pengiriman","tanggal_pengiriman","FOB","syarat_pembayaran"]]
        
        # EXPENSE rows
        expense_rows = data[["rank"]].drop_duplicates().copy()
        expense_rows["row_or_header"] = 2
        expense_rows["rank_part"] = 3
        expense_rows["header"] = "EXPENSE"
        expense_rows["no_form"] = 0
        expense_rows["tgl_pesanan"] = ""
        expense_rows["no_pelanggan"] = ""
        expense_rows["no_po"] = ""
        expense_rows["alamat"] = ""
        expense_rows["kena_ppn"] = ""
        expense_rows["total_termasuk_ppn"] = ""
        expense_rows["diskon_pesanan_percentage"] = ""
        expense_rows["diskon_pesanan_rupiah"] = ""
        expense_rows["keterangan"] = ""
        expense_rows["nama_cabang"] = ""
        expense_rows["pengiriman"] = ""
        expense_rows["tanggal_pengiriman"] = ""
        expense_rows["FOB"] = ""
        expense_rows["syarat_pembayaran"] = ""
        expense_rows = expense_rows[["row_or_header", "rank", "rank_part", "header", "no_form", "tgl_pesanan", "no_pelanggan","no_po","alamat","kena_ppn","total_termasuk_ppn","diskon_pesanan_percentage","diskon_pesanan_rupiah","keterangan","nama_cabang","pengiriman","tanggal_pengiriman","FOB","syarat_pembayaran"]]

        transform_data = pd.concat([header_rows, item_rows, expense_rows], ignore_index=True)
        transform_data = transform_data.sort_values(by=["rank", "rank_part"]).reset_index(drop=True)

        # Load Excel template
        excel_path = "header.xlsx"  # Adjust this path to your template
        wb = load_workbook(excel_path)
        ws = wb.active

        start_row = 4
        for r_idx, row in enumerate(dataframe_to_rows(transform_data.iloc[:, 3:], index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"final_sales_order_{timestamp}.xlsx"

    st.success("Processing complete!")

    st.download_button(
        label="Download",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )






